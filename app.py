import io
from datetime import date, timedelta

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

import config
import data.database as db
from core.buff_scraper import (
    diagnose_price_extraction, ensure_login, get_full_price_history,
    get_items_on_date, get_price_history, is_logged_in, open_buff_page,
)
from core.steam_scraper import (
    diagnose_steam_extraction, ensure_steam_login,
    get_steam_market_data, is_steam_logged_in, open_steam_market,
)
import core.steam_scraper as _steam
from core.filters import apply_initial_filters, is_price_stable
from data.models import ItemSnapshot, PriceRecord
from utils.helpers import sleep_random

st.set_page_config(page_title="BUFF数据提取", layout="wide")
db.init_db()

# 隐藏 Streamlit 默认 UI 元素（Deploy 按钮 / Made with Streamlit 等英文组件）
st.markdown("""
<style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

st.title("BUFF数据提取")
st.caption("版本: 2026-05-26-v10 | 一键获取+分步执行双层设计 | BUFF筛选→稳定性→Steam数据→套利对比")

# ---------- 侧边栏参数配置 ----------
with st.sidebar:
    st.header("参数配置")

    target_date = st.date_input("目标日期", value=config.DEFAULT_TARGET_DATE)
    stable_days = st.number_input(
        "价格稳定考察天数", min_value=1, max_value=90, value=config.DEFAULT_STABLE_DAYS, key="stable_days_v2"
    )
    volatility_threshold = st.slider(
        "价格波动阈值 (%)", min_value=1, max_value=30, value=5
    ) / 100.0

    st.divider()
    st.subheader("卡价转换比")
    conversion_rate = st.number_input(
        "卡价转换比（美元 → 人民币）",
        min_value=0.1, max_value=20.0, value=5.0, step=0.01,
        help="用于将 Steam 美元价格转换为人民币，与 BUFF 价格进行对比",
    )

    st.divider()
    st.subheader("BUFF 筛选条件")
    category_names = st.multiselect(
        "饰品品类",
        options=list(config.CATEGORY_OPTIONS.keys()),
        default=["全部/不限"],
        help="可多选。选「全部/不限」或全不选时不做品类限制；选多个品类时各品类独立抓取，数量均分",
    )
    target_count = st.number_input(
        "目标饰品数量", min_value=10, max_value=2000, value=config.DEFAULT_TARGET_COUNT, step=10,
        help="BUFF 端初步筛选的目标数量，不含 Steam 稳定性筛选",
    )
    min_price = st.number_input(
        "最低价格", min_value=0.0, max_value=10000.0, value=float(config.DEFAULT_MIN_PRICE), step=1.0,
        help="BUFF 在售最低价格（元），低于此价格的饰品被过滤",
    )
    min_volume = st.number_input(
        "最低在售数量", min_value=1, max_value=100000, value=config.DEFAULT_MIN_VOLUME, step=10,
        help="BUFF 在售数量下限，低于此数量的饰品被过滤",
    )

    st.divider()
    if is_logged_in():
        if st.button("查看 BUFF 页面", use_container_width=True):
            with st.spinner("正在打开 BUFF 页面..."):
                open_buff_page()
            st.success("浏览器窗口已关闭")
    else:
        if st.button("登录 BUFF", use_container_width=True):
            with st.spinner("正在打开浏览器，请完成登录..."):
                ensure_login()
            if is_logged_in():
                st.success("登录态已保存")
            else:
                st.error("未检测到登录态，请重新尝试")

    st.divider()
    if is_steam_logged_in():
        if st.button("查看 Steam 市场", use_container_width=True):
            with st.spinner("正在打开 Steam 市场..."):
                open_steam_market()
            st.success("浏览器窗口已关闭")
    else:
        if st.button("登录 Steam 账号", use_container_width=True):
            with st.spinner("正在打开浏览器，请完成 Steam 登录..."):
                ensure_steam_login()
            if is_steam_logged_in():
                st.success("Steam 登录态已保存")
            else:
                st.error("未检测到 Steam 登录态，请重新尝试")

    st.divider()
    st.caption(
        f"配置预览：\n"
        f"- 目标日期：{target_date}\n"
        f"- 考察 {stable_days} 天价格波动（≤{volatility_threshold * 100:.0f}%）\n"
        f"- 卡价转换比（USD→CNY）：{conversion_rate}\n"
        f"- 品类：{'、'.join(category_names) if category_names else '全部/不限'}\n"
        f"- 目标 {target_count} 件 | 最低 ¥{min_price} | 在售≥{min_volume}"
    )

    # 分步模式参数变更检测
    _snap = st.session_state.get("_param_snapshot")
    if _snap and st.session_state.get("one_click_mode") is None:
        affected = []
        if st.session_state.get("stage1_done"):
            if (category_names != _snap.get("category_names") or
                target_count != _snap.get("target_count") or
                min_price != _snap.get("min_price") or
                min_volume != _snap.get("min_volume")):
                affected.append("Step ① 品类/数量/最低价/在售量")
        if st.session_state.get("stage2_done"):
            if (str(target_date) != _snap.get("target_date") or
                stable_days != _snap.get("stable_days") or
                volatility_threshold != _snap.get("volatility_threshold")):
                affected.append("Step ② 目标日期/考察天数/波动阈值")
        if st.session_state.get("stage4_done"):
            if conversion_rate != _snap.get("conversion_rate"):
                affected.append("Step ④ 汇率转换比")
        if affected:
            st.warning("参数已修改，影响已完成步骤：\n" + "\n".join(f"• {a}" for a in affected) +
                       "\n建议重新执行受影响的步骤")


# ---------- 一键获取：执行全部四步 ----------
def _log_error(step: int, item_id: str, item_name: str, error: str,
               context: dict = None):
    """记录一条错误到 session_state.error_log，含可选的现场上下文。"""
    from datetime import datetime as _dt
    entry = {
        "step": step,
        "item_id": item_id,
        "item_name": item_name[:60] if item_name else "",
        "error": error,
        "time": _dt.now().strftime("%H:%M:%S"),
    }
    if context:
        entry["context"] = context
    st.session_state.error_log.append(entry)


def _clear_downstream_steps(from_step: int):
    """清空从 from_step 开始的所有下游结果（1-indexed）。"""
    if from_step <= 1:
        st.session_state.stage1_done = False
        st.session_state.raw_items = []
        st.session_state.filtered_items = []
    if from_step <= 2:
        st.session_state.stage2_done = False
        st.session_state.stable_items = []
    if from_step <= 3:
        st.session_state.stage3_done = False
        st.session_state.steam_data = {}
    if from_step <= 4:
        st.session_state.stage4_done = False
        st.session_state.arbitrage_results = {}


def _show_error_log():
    """在页面上展示错误日志 expander，含可展开的上下文细节。"""
    errors = st.session_state.get("error_log", [])
    if not errors:
        return
    step_names = {1: "BUFF筛选", 2: "稳定性筛选", 3: "Steam数据", 4: "套利对比"}
    with st.expander(f"⚠️ 错误日志（{len(errors)} 条）", expanded=len(errors) > 0):
        for e in errors:
            icon = "❌"
            step_label = step_names.get(e["step"], f"Step{e['step']}")
            item_label = f" [{e['item_name']}]" if e["item_name"] else ""
            ctx = e.get("context")
            ctx_summary = ""
            if ctx:
                etype = ctx.get("type", "")
                detail = ctx.get("detail", "")
                ctx_summary = f" [{etype}] {detail[:60]}"
            st.caption(
                f"{icon} **Step{e['step']}·{step_label}**{item_label}"
                f" — {e['error']}{ctx_summary}  `{e['time']}`"
            )
            # 展开上下文详情
            if ctx:
                with st.status("", expanded=False):
                    for k, v in ctx.items():
                        val = str(v)
                        if len(val) > 300:
                            val = val[:300] + "..."
                        st.code(f"{k}: {val}")

        # 导出诊断信息
        import json as _json
        _diag = []
        for e in errors:
            _diag.append({
                "step": e["step"], "item_id": e.get("item_id", ""),
                "item_name": e.get("item_name", ""), "error": e["error"],
                "context": e.get("context", {}),
            })
        _diag_json = _json.dumps(_diag, ensure_ascii=False, indent=2,
                                 default=str)
        st.download_button(
            "📋 导出诊断信息",
            data=_diag_json,
            file_name=f"error_diagnostic_{date.today().isoformat()}.json",
            mime="application/json",
            use_container_width=True,
            help="导出所有错误日志及上下文，供开发者分析",
        )


def _snapshot_params():
    """保存当前 sidebar 参数快照，用于分步模式下检测参数变更。"""
    st.session_state._param_snapshot = {
        "category_names": list(category_names) if category_names else [],
        "target_count": target_count,
        "min_price": min_price,
        "min_volume": min_volume,
        "target_date": str(target_date),
        "stable_days": stable_days,
        "volatility_threshold": volatility_threshold,
        "conversion_rate": conversion_rate,
    }


def _execute_step1(target_date, target_count, categories, min_price, min_volume, run_id=None):
    try:
        st.session_state.raw_items = get_items_on_date(
            target_date, target_count=target_count,
            categories=categories, min_price=min_price, min_volume=min_volume)
    except Exception as e:
        _log_error(1, "", "全部", f"BUFF列表获取异常: {e}")
        st.session_state.raw_items = []
    st.session_state.filtered_items = apply_initial_filters(
        st.session_state.raw_items, min_price=min_price, min_volume=min_volume)
    st.session_state.stage1_done = True
    if run_id:
        db.save_step1(run_id, st.session_state.raw_items, st.session_state.filtered_items)
    if not st.session_state.filtered_items:
        _log_error(1, "", "全部", f"初步筛选后无饰品通过（原始{len(st.session_state.raw_items)}条，在售>100且价格>20）")


def _execute_step2(target_date, stable_days, volatility_threshold, run_id=None):
    filtered = st.session_state.filtered_items
    stable = []
    total = len(filtered)
    progress_bar = st.progress(0, text="准备开始...")
    try:
        for idx, item in enumerate(filtered):
            progress_bar.progress((idx) / total, text=f"正在获取 ({idx+1}/{total}) {item.name[:40]}…")
            st.write(f"  [{idx+1}/{total}] 正在获取 {item.name} 的价格历史…")
            start = target_date - timedelta(days=stable_days)
            history = get_price_history(item.item_id, start, target_date)
            item._debug_history_len = len(history)
            if history:
                prices = [r.price for r in history]
                item._debug_min_price = min(prices)
                item._debug_max_price = max(prices)
                item._debug_volatility = (max(prices) - min(prices)) / (sum(prices)/len(prices)) if sum(prices) > 0 else 0
                st.write(f"    价格记录: {len(history)} 条, 范围 ¥{min(prices):.2f} ~ ¥{max(prices):.2f}, "
                         f"波动 {item._debug_volatility*100:.1f}%")
                for r in history[:5]:
                    st.write(f"      · {r.date.isoformat()}: ¥{r.price:.2f}")
                if len(history) > 5:
                    st.write(f"      ... 及 {len(history)-5} 条更多记录")
            if is_price_stable(history, volatility_threshold):
                item.price_history = history
                stable.append(item)
                st.write(f"    ✅ 通过 (波动 {item._debug_volatility*100:.1f}% ≤ {volatility_threshold*100:.0f}%)")
            else:
                if len(history) == 0:
                    item._debug_fail_reason = "无价格数据"
                    _log_error(2, item.item_id, item.name, "BUFF价格历史为空")
                    st.write(f"    ❌ 未通过: 无价格数据")
                else:
                    vol = item._debug_volatility
                    item._debug_fail_reason = f"波动 {vol*100:.1f}% > {volatility_threshold*100:.0f}%"
                    _log_error(2, item.item_id, item.name, item._debug_fail_reason)
                    st.write(f"    ❌ 未通过: 波动 {vol*100:.1f}% > 阈值 {volatility_threshold*100:.0f}%")
    except Exception as e:
        import traceback
        _log_error(2, "", "全部", f"Step 2 执行异常（已保留部分数据）: {e}")
        st.write(f"  ❌ **Step 2 执行中断**: {e}，已处理 {len(stable)}/{total} 个饰品")
        traceback.print_exc()
    progress_bar.progress(1.0, text="完成!")
    progress_bar.empty()
    st.session_state.stable_items = stable
    st.session_state.stage2_done = True
    if run_id:
        db.save_step2(run_id, filtered, stable)


def _execute_step3(run_id=None, status=None):
    stable = st.session_state.stable_items
    steam_data = {}
    st.session_state.failed_step3_items = []

    try:
        # 按基础皮肤名分组，同组共用一次 Steam 页面调度
        groups = _steam.group_by_skin_name(stable)
        group_count = len(groups)
        group_idx = 0

        for base_name, group_items in groups.items():
            group_idx += 1
            if status:
                status.update(label=f"第3步/共4步：Steam市场数据获取… [{group_idx}/{group_count}] 组")
            st.write(f"  ── [{group_idx}/{group_count}] 皮肤组: {base_name} "
                     f"（{len(group_items)} 个变体）──")

            # 准备批处理输入
            group_members = []
            for item in group_items:
                target_dates = sorted(set(
                    r.date - timedelta(days=7) for r in item.price_history
                ))
                group_members.append({
                    "item_id": item.item_id,
                    "buff_item_name": item.name,
                    "target_dates": target_dates,
                    "base_skin_name": base_name,
                })

            # 批处理：一次 BUFF + 一次 Steam 获取所有变体数据
            batch_results = _steam.get_steam_market_data_batch(
                representative_item_id=group_items[0].item_id,
                group_members=group_members,
            )

            # 更新每个变体的数据
            for item in group_items:
                data = batch_results.get(item.item_id)
                if data and data.get("steam_price_history"):
                    steam_data[item.item_id] = data
                    item.steam_url = data.get("steam_url")
                    item.steam_price = data.get("steam_price")
                    item.steam_sold_count = data.get("steam_sold_count", 0)
                    item.steam_price_history = data.get("steam_price_history", [])
                    steam_price_str = f"${item.steam_price:.2f}" if item.steam_price else "N/A"
                    st.write(f"    ✅ {item.name}: Steam {steam_price_str}, "
                             f"售出 {item.steam_sold_count} 件")
                    if item.steam_price_history:
                        st.write(f"    价格历史:")
                        for r in item.steam_price_history[:10]:
                            st.write(f"      · {r.date.isoformat()}: ${r.price:.2f}, "
                                     f"销量 {r.volume} 件")
                        if len(item.steam_price_history) > 10:
                            st.write(f"      ... 及 {len(item.steam_price_history)-10} 条")
                    date_records = data.get("date_records", [])
                    if date_records:
                        st.write(f"    各日期节点:")
                        for dr in date_records:
                            st.write(f"      · {dr['date']}: ${dr['steam_price']:.2f}, "
                                     f"销量 {dr['steam_volume']} 件")
                else:
                    reason = _steam.get_last_steam_error(item_id=item.item_id)
                    if data and not data.get("steam_price_history"):
                        reason = reason or "Steam API 返回无价格数据"
                    else:
                        reason = reason or "Steam数据获取失败"
                    ctx = _steam.get_last_steam_error_context(item_id=item.item_id)
                    _log_error(3, item.item_id, item.name, reason, context=ctx)
                    st.write(f"    ❌ {item.name}: 获取失败 - {reason}")
                    st.session_state.failed_step3_items.append({
                        "item_id": item.item_id,
                        "buff_item_name": item.name,
                        "target_dates": [
                            r.date - timedelta(days=7) for r in item.price_history
                        ],
                        "base_skin_name": base_name,
                    })

            if group_idx < group_count:
                sleep_random(2.0, 4.0)

    except Exception as e:
        import traceback
        _log_error(3, "", "全部", f"Step 3 执行异常（已保留部分数据）: {e}")
        st.write(f"  ❌ **Step 3 执行中断**: {e}")
        traceback.print_exc()

    st.session_state.steam_data = steam_data
    st.session_state.stage3_done = True
    if run_id:
        db.save_step3(run_id, stable, steam_data)


def _execute_step4(conversion_rate, run_id=None):
    stable = st.session_state.stable_items
    steam_data = st.session_state.steam_data
    arbitrage_results = {}
    try:
        for item in stable:
            data = steam_data.get(item.item_id)
            if not data or not data.get("steam_price_history"):
                _log_error(4, item.item_id, item.name, "无Steam价格历史，跳过套利对比")
                continue
            buff_history = sorted(item.price_history, key=lambda r: r.date)
            steam_history = sorted(data["steam_price_history"], key=lambda r: r.date)
            buff_by_date = {r.date: r.price for r in buff_history}
            steam_by_buff_date = {}
            for r in steam_history:
                steam_by_buff_date[r.date + timedelta(days=7)] = r
            date_pairs = []
            for buff_date in sorted(buff_by_date.keys()):
                sr = steam_by_buff_date.get(buff_date)
                if sr:
                    steam_price_cny = sr.price * conversion_rate
                    diff = buff_by_date[buff_date] - steam_price_cny
                    date_pairs.append({
                        "buff_date": buff_date,
                        "buff_price": buff_by_date[buff_date],
                        "steam_date": sr.date,
                        "steam_price_usd": sr.price,
                        "steam_price_cny": steam_price_cny,
                        "steam_volume": sr.volume,
                        "diff": diff,
                        "is_target": diff > 0,
                    })
            if date_pairs:
                avg_buff = sum(p["buff_price"] for p in date_pairs) / len(date_pairs)
                avg_steam_usd = sum(p["steam_price_usd"] for p in date_pairs) / len(date_pairs)
                avg_steam_cny = avg_steam_usd * conversion_rate
                avg_diff = avg_buff - avg_steam_cny
                is_target = avg_diff > 0
                arbitrage_results[item.item_id] = {
                    "date_pairs": date_pairs,
                    "avg_buff_price": avg_buff,
                    "avg_steam_usd": avg_steam_usd,
                    "avg_steam_cny": avg_steam_cny,
                    "avg_diff": avg_diff,
                    "is_target": is_target,
                    "target_count": sum(1 for p in date_pairs if p["is_target"]),
                }
    except Exception as e:
        import traceback
        _log_error(4, "", "全部", f"Step 4 执行异常: {e}")
        st.write(f"  ❌ **Step 4 执行中断**: {e}")
        traceback.print_exc()
    st.session_state.arbitrage_results = arbitrage_results
    st.session_state.stage4_done = True
    st.session_state._run_conversion_rate = conversion_rate
    if run_id:
        db.save_step4(run_id, stable, arbitrage_results)
        db.finish_run(run_id)


# ---------- 主区域 ----------
tabs = st.tabs(["选品流程", "价格走势查询"])

with tabs[0]:
    # ---- 初始化 session_state ----
    defaults = {
        "stage1_done": False, "raw_items": [], "filtered_items": [],
        "stage2_done": False, "stable_items": [],
        "stage3_done": False, "steam_data": {},
        "stage4_done": False, "arbitrage_results": {},
        "one_click_mode": None,  # None | "running" | "done"
        "current_run_id": None,  # DB run id for persistence
        "error_log": [],  # [{step, item_id, item_name, error, time}]
        "_param_snapshot": None,
        "_run_conversion_rate": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    one_click_mode = st.session_state.one_click_mode

    # =====================================================================
    # 一键获取完成 → 展示结果 + 执行细则
    # =====================================================================
    if one_click_mode == "done" and st.session_state.stage4_done:
        arbitrage_results = st.session_state.arbitrage_results
        stable = st.session_state.stable_items
        steam_data = st.session_state.steam_data
        filtered = st.session_state.filtered_items
        raw = st.session_state.raw_items
        _disp_rate = st.session_state.get("_run_conversion_rate", conversion_rate)

        st.success(f"一键获取完成！BUFF 原始 {len(raw)} 条 → 初步过滤 {len(filtered)} 条 "
                   f"→ 价格稳定 {len(stable)} 条 → Steam 数据 {len(steam_data)} 条")

        # ── 分步状态 & 重新执行按钮 ──
        st.divider()
        _run_id_redo = st.session_state.current_run_id
        s1_raw = len(raw)
        s1_ok = len(filtered)
        s2_ok = len(stable)
        s3_ok = len(steam_data)
        s3_total = len(stable)
        s4_ok = sum(1 for v in arbitrage_results.values() if v.get("is_target")) if arbitrage_results else 0
        s4_total = len(arbitrage_results)

        col_r1, col_r2, col_r3, col_r4 = st.columns(4)

        with col_r1:
            st.metric("Step 1 BUFF筛选", f"{s1_ok} 条", delta=f"原始{s1_raw}条")
            if st.button("🔄 重新执行 Step 1→4", key="redo_s1", use_container_width=True):
                _cat_vals = [config.CATEGORY_OPTIONS[n] for n in category_names if n != "全部/不限"]
                _clear_downstream_steps(1)
                _execute_step1(target_date, target_count, _cat_vals, min_price, min_volume, _run_id_redo)
                if st.session_state.filtered_items:
                    _execute_step2(target_date, stable_days, volatility_threshold, _run_id_redo)
                if st.session_state.stable_items:
                    _execute_step3(_run_id_redo)
                if st.session_state.steam_data:
                    _execute_step4(conversion_rate, _run_id_redo)
                st.rerun()

        with col_r2:
            _pass_rate = f"{s2_ok}/{len(filtered)}" if filtered else "N/A"
            st.metric("Step 2 稳定性", f"{s2_ok} 条", delta=_pass_rate)
            if st.button("🔄 重新执行 Step 2→4", key="redo_s2", use_container_width=True):
                _clear_downstream_steps(2)
                _execute_step2(target_date, stable_days, volatility_threshold, _run_id_redo)
                if st.session_state.stable_items:
                    _execute_step3(_run_id_redo)
                if st.session_state.steam_data:
                    _execute_step4(conversion_rate, _run_id_redo)
                st.rerun()

        with col_r3:
            _s3_rate = f"{s3_ok}/{s3_total}" if s3_total else "N/A"
            _s3_delta = "全部成功 ✅" if s3_ok >= s3_total else f"{s3_total-s3_ok}条失败 ⚠️" if s3_total else "N/A"
            st.metric("Step 3 Steam数据", _s3_rate, delta=_s3_delta)
            if st.button("🔄 重新执行 Step 3→4", key="redo_s3", use_container_width=True):
                _clear_downstream_steps(3)
                _execute_step3(_run_id_redo)
                if st.session_state.steam_data:
                    _execute_step4(conversion_rate, _run_id_redo)
                st.rerun()

        with col_r4:
            _s4_delta = f"共{s4_total}个对比" if s4_total else "无数据 ❌"
            st.metric("Step 4 套利结果", f"{s4_ok} 个目标", delta=_s4_delta)
            if st.button("🔄 重新执行 Step 4", key="redo_s4", use_container_width=True):
                _clear_downstream_steps(4)
                _execute_step4(conversion_rate, _run_id_redo)
                st.rerun()

        # ── 失败饰品重试 ──
        _failed_step3 = st.session_state.get("failed_step3_items", [])
        if _failed_step3:
            st.divider()
            with st.expander(f"🔄 重试失败饰品（{len(_failed_step3)} 条）", expanded=False):
                st.markdown("以下饰品 Step 3 获取失败，可单独或批量重试：")
                for fi in _failed_step3:
                    name = fi["buff_item_name"]
                    col_a, col_b = st.columns([4, 1])
                    with col_a:
                        st.caption(f"{name}  (ID: {fi['item_id']})")
                    with col_b:
                        retry_key = f"redo_retry_{fi['item_id']}"
                        if st.button("重试", key=retry_key, use_container_width=True):
                            with st.spinner(f"正在重试 {name}..."):
                                result = _steam.get_steam_market_data(
                                    fi["item_id"], fi["target_dates"], fi["buff_item_name"],
                                )
                            if result:
                                st.session_state.steam_data[fi["item_id"]] = result
                                st.session_state.failed_step3_items = [
                                    f for f in st.session_state.failed_step3_items
                                    if f["item_id"] != fi["item_id"]
                                ]
                                st.success(f"✅ {name} 重试成功，正在重新计算套利对比...")
                                _execute_step4(conversion_rate, _run_id_redo)
                                st.rerun()
                            else:
                                err = _steam.get_last_steam_error(item_id=fi["item_id"]) or "未知错误"
                                _log_error(3, fi["item_id"], name, err)
                                st.error(f"❌ {name} 重试失败: {err}")

                st.divider()
                if st.button("🔄 重试全部失败饰品", type="primary",
                              use_container_width=True, key="redo_retry_all"):
                    with st.status("正在批量重试所有失败饰品...", expanded=True) as retry_status:
                        batch_result = _steam.retry_steam_failed_items(_failed_step3)
                        ok = 0
                        for fi in _failed_step3:
                            data = batch_result.get(fi["item_id"])
                            if data:
                                st.session_state.steam_data[fi["item_id"]] = data
                                ok += 1
                                retry_status.write(f"✅ {fi['buff_item_name']} 成功")
                            else:
                                err = _steam.get_last_steam_error(item_id=fi["item_id"]) or "失败"
                                _log_error(3, fi["item_id"], fi["buff_item_name"], err)
                                retry_status.write(f"❌ {fi['buff_item_name']} 失败")
                        st.session_state.failed_step3_items = [
                            f for f in _failed_step3
                            if f["item_id"] not in batch_result
                        ]
                        st.success(f"重试完成: {ok}/{len(_failed_step3)} 成功，正在重新计算套利对比...")
                        _execute_step4(conversion_rate, _run_id_redo)
                        st.rerun()

        st.divider()

        if arbitrage_results:
            target_items = [it for it in stable if arbitrage_results.get(it.item_id, {}).get("is_target")]
            st.header(f"🎯 发现 {len(target_items)} 个目标饰品")

            # ---- 导出Excel ----
            if target_items:
                buf = io.BytesIO()
                export_rows = []
                for item in target_items:
                    ar = arbitrage_results.get(item.item_id, {})
                    _steam_url = steam_data.get(item.item_id, {}).get("steam_url", "")
                    for p in ar.get("date_pairs", []):
                        if p["is_target"]:
                            export_rows.append({
                                "饰品名称": item.name,
                                "BUFF日期": str(p["buff_date"]),
                                "BUFF价格(¥)": round(p["buff_price"], 2),
                                "Steam日期": str(p["steam_date"]),
                                "Steam价格($)": round(p["steam_price_usd"], 2),
                            })
                # 目标饰品列表（含 Steam 链接）
                target_rows = []
                for item in target_items:
                    ar = arbitrage_results.get(item.item_id, {})
                    _steam_url = steam_data.get(item.item_id, {}).get("steam_url", "")
                    target_rows.append({
                        "饰品名称": item.name,
                        "BUFF均价(¥)": f"¥{ar['avg_buff_price']:.2f}",
                        "Steam均价(¥)": f"¥{ar['avg_steam_cny']:.2f}",
                        "均价差(¥)": f"¥{ar['avg_diff']:+.2f}",
                        "Steam链接": _steam_url if _steam_url else "无",
                    })
                st.dataframe(
                    pd.DataFrame(target_rows),
                    column_config={
                        "Steam链接": st.column_config.LinkColumn("Steam链接", display_text="打开"),
                    },
                    use_container_width=True,
                )
                if export_rows:
                    df_export = pd.DataFrame(export_rows)
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils import get_column_letter
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        df_export.to_excel(writer, index=False, sheet_name="目标饰品")
                        ws = writer.sheets["目标饰品"]
                        # 表头加粗 + 浅灰底
                        header_font = Font(bold=True)
                        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        for cell in ws[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center")
                        # 列宽自适应
                        for col_idx, col in enumerate(df_export.columns, 1):
                            max_len = max(
                                df_export[col].astype(str).map(len).max(),
                                len(str(col)),
                            )
                            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)
                        # 冻结首行
                        ws.freeze_panes = "A2"
                        # 价格列数字格式
                        for col_idx, col in enumerate(df_export.columns, 1):
                            if "价格" in col:
                                for row in range(2, ws.max_row + 1):
                                    ws.cell(row=row, column=col_idx).number_format = '#,##0.00'
                    _target_date_str = str(target_date)
                    st.download_button(
                        label="📥 导出Excel",
                        data=buf.getvalue(),
                        file_name=f"套利结果_{_target_date_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

            # ---- 主结果表格 ----
            rows = []
            for item in stable:
                ar = arbitrage_results.get(item.item_id)
                if ar:
                    date_details = "; ".join(
                        f"{p['buff_date']}: BUFF ¥{p['buff_price']:.2f} vs "
                        f"Steam ${p['steam_price_usd']:.2f}×{_disp_rate}=¥{p['steam_price_cny']:.2f} "
                        f"→ {'✅' if p['is_target'] else '❌'}差¥{p['diff']:+.2f}"
                        for p in ar["date_pairs"]
                    )
                    _steam_url = steam_data.get(item.item_id, {}).get("steam_url", "")
                    rows.append({
                        "饰品ID": item.item_id,
                        "名称": item.name,
                        "BUFF均价(¥)": f"{ar['avg_buff_price']:.2f}",
                        "Steam均价($)": f"${ar['avg_steam_usd']:.2f}",
                        "Steam均价(¥)": f"¥{ar['avg_steam_cny']:.2f}",
                        "均价差(¥)": f"¥{ar['avg_diff']:+.2f}",
                        "命中节点": f"{ar['target_count']}/{len(ar['date_pairs'])}",
                        "判定": "🎯 目标" if ar["is_target"] else "未达标",
                        "各节点明细": date_details,
                        "Steam链接": _steam_url if _steam_url else "无",
                    })
                else:
                    _steam_url = steam_data.get(item.item_id, {}).get("steam_url", "")
                    rows.append({
                        "饰品ID": item.item_id,
                        "名称": item.name,
                        "BUFF均价(¥)": item.buff_price,
                        "Steam均价($)": "无数据",
                        "Steam均价(¥)": "无数据",
                        "均价差(¥)": "N/A",
                        "命中节点": "N/A",
                        "判定": "无Steam数据",
                        "各节点明细": "无",
                        "Steam链接": _steam_url if _steam_url else "无",
                    })
            st.dataframe(
                pd.DataFrame(rows),
                column_config={
                    "Steam链接": st.column_config.LinkColumn("Steam链接", display_text="打开"),
                },
                use_container_width=True,
            )

        # ---- 查看执行细则 ----
        with st.expander("📋 查看执行细则", expanded=False):
            st.subheader("第一步：BUFF初步筛选")
            st.write(f"BUFF 原始获取：{len(raw)} 条 → 初步过滤后：{len(filtered)} 条")
            if filtered:
                df1 = pd.DataFrame([{
                    "饰品ID": it.item_id, "名称": it.name,
                    "BUFF价格": it.buff_price, "在售数量": it.volume, "成交额": it.turnover,
                } for it in filtered])
                st.dataframe(df1, use_container_width=True)

            st.divider()
            st.subheader("第二步：价格稳定性筛选")
            st.write(f"价格稳定（波动 ≤ {volatility_threshold * 100:.0f}%）的饰品：{len(stable)} / {len(filtered)} 条")
            if filtered:
                detail_rows = []
                for item in filtered:
                    hist_len = getattr(item, '_debug_history_len', -1)
                    min_p = getattr(item, '_debug_min_price', 0)
                    max_p = getattr(item, '_debug_max_price', 0)
                    vol = getattr(item, '_debug_volatility', 0)
                    passed = item in stable
                    fail_reason = getattr(item, '_debug_fail_reason', '')
                    detail_rows.append({
                        "饰品": item.name,
                        "价格记录数": hist_len,
                        "最低": f"{min_p:.2f}" if hist_len > 0 else "N/A",
                        "最高": f"{max_p:.2f}" if hist_len > 0 else "N/A",
                        "波动": f"{vol*100:.1f}%" if hist_len > 0 else "N/A",
                        "结果": "✅" if passed else f"❌ {fail_reason}",
                    })
                st.dataframe(pd.DataFrame(detail_rows), use_container_width=True)

            st.divider()
            st.subheader("第三步：Steam市场数据")
            st.write(f"成功获取Steam市场数据：{len(steam_data)} / {len(stable)} 条")
            if steam_data:
                s3_rows = []
                for item in stable:
                    data = steam_data.get(item.item_id)
                    if data:
                        date_records = data.get("date_records", [])
                        date_details = "; ".join(
                            f"{d['date']}: ${d['steam_price']:.2f}/{d['steam_volume']}件"
                            for d in date_records
                        ) if date_records else "无数据"
                        sp = data.get("steam_price")
                        sc = data.get("steam_sold_count", 0)
                        steam_price_str = f"${sp:.2f}" if sp else "N/A"
                        s3_rows.append({
                            "饰品ID": item.item_id, "名称": item.name,
                            "BUFF价格": item.buff_price, "Steam均价": steam_price_str,
                            "总售出量": sc, "各日期节点明细": date_details,
                            "Steam链接": data.get("steam_url", "获取失败")[:80] + "...",
                        })
                    else:
                        s3_rows.append({
                            "饰品ID": item.item_id, "名称": item.name,
                            "BUFF价格": item.buff_price, "Steam均价": "获取失败",
                            "总售出量": 0, "各日期节点明细": "获取失败",
                            "Steam链接": "获取失败",
                        })
                st.dataframe(pd.DataFrame(s3_rows), use_container_width=True)

            st.divider()
            st.subheader("第四步：套利对比")
            if arbitrage_results:
                ar_rows = []
                for item in stable:
                    ar = arbitrage_results.get(item.item_id)
                    if ar:
                        ar_rows.append({
                            "饰品": item.name,
                            "对比节点数": len(ar["date_pairs"]),
                            "BUFF均价": f"¥{ar['avg_buff_price']:.2f}",
                            "Steam转换均价": f"¥{ar['avg_steam_cny']:.2f}",
                            "均价差": f"¥{ar['avg_diff']:+.2f}",
                            "命中/总节点": f"{ar['target_count']}/{len(ar['date_pairs'])}",
                            "结果": "🎯 目标" if ar["is_target"] else "未达标",
                        })
                    else:
                        ar_rows.append({
                            "饰品": item.name, "对比节点数": 0,
                            "BUFF均价": f"{item.buff_price:.2f}",
                            "Steam转换均价": "N/A", "均价差": "N/A",
                            "命中/总节点": "N/A", "结果": "❌ 无Steam数据",
                        })
                st.dataframe(pd.DataFrame(ar_rows), use_container_width=True)

                # Date pair details
                with st.expander("📊 按日期节点查看配对对比"):
                    for item in stable:
                        ar = arbitrage_results.get(item.item_id)
                        if ar and ar["date_pairs"]:
                            st.markdown(f"**{item.name}** (ID: {item.item_id})  "
                                        f"{'🎯 目标' if ar['is_target'] else '未达标'}")
                            pair_rows = []
                            for p in ar["date_pairs"]:
                                pair_rows.append({
                                    "BUFF日期": p["buff_date"], "BUFF价格(¥)": f"{p['buff_price']:.2f}",
                                    "Steam日期": p["steam_date"], "Steam价格($)": f"{p['steam_price_usd']:.2f}",
                                    "Steam转换(¥)": f"{p['steam_price_cny']:.2f}",
                                    "差价(¥)": f"{p['diff']:+.2f}", "Steam销量": p["steam_volume"],
                                    "判定": "✅ 目标" if p["is_target"] else "❌",
                                })
                            st.dataframe(pd.DataFrame(pair_rows), use_container_width=True)
                            st.divider()

                # Charts
                with st.expander("📈 BUFF vs Steam 价格对比图表"):
                    seen = {}
                    chart_idx = 0
                    for item in stable:
                        if item.item_id in seen:
                            continue
                        seen[item.item_id] = True
                        data = steam_data.get(item.item_id)
                        ar = arbitrage_results.get(item.item_id)
                        if data and data.get("steam_price_history"):
                            try:
                                fig = go.Figure()
                                if item.price_history:
                                    buff_dates = [r.date for r in item.price_history]
                                    buff_prices = [r.price for r in item.price_history]
                                    fig.add_trace(go.Scatter(
                                        x=buff_dates, y=buff_prices, name=f"{item.name} BUFF(¥)",
                                        mode="lines+markers", yaxis="y",
                                    ))
                                steam_hist = data["steam_price_history"]
                                steam_dates = [r.date for r in steam_hist]
                                steam_prices_cny = [r.price * _disp_rate for r in steam_hist]
                                fig.add_trace(go.Scatter(
                                    x=steam_dates, y=steam_prices_cny, name=f"{item.name} Steam→¥",
                                    mode="lines+markers", yaxis="y",
                                ))
                                title_suffix = " 🎯目标" if (ar and ar["is_target"]) else ""
                                fig.update_layout(
                                    title=f"{item.name} 价格对比 (1 USD = {_disp_rate} CNY){title_suffix}",
                                    xaxis_title="日期", yaxis_title="价格 (¥)",
                                    hovermode="x unified",
                                )
                                st.plotly_chart(fig, use_container_width=True,
                                                key=f"oneclick_chart_{chart_idx}")
                                chart_idx += 1
                            except ImportError:
                                st.info("安装 plotly 可查看价格对比图表：pip install plotly")
                                break

        # Reset
        _show_error_log()
        st.divider()
        if st.button("🔄 重新开始（清除所有结果）", use_container_width=True):
            for k in defaults:
                st.session_state[k] = defaults[k]
            st.rerun()

    # =====================================================================
    # 未完成一键获取 → 显示一键按钮 + 分步执行
    # =====================================================================
    else:
        # ---- 一键获取按钮 ----
        st.header("🚀 一键获取选品")
        st.caption("自动依次执行：BUFF初步筛选 → 价格稳定性分析 → Steam市场数据获取 → 套利对比，完成后展示结果。")

        if st.button("开始一键获取", type="primary", use_container_width=True):
            st.session_state.one_click_mode = "running"
            st.session_state.error_log = []
            category_values = [config.CATEGORY_OPTIONS[n] for n in category_names if n != "全部/不限"]
            _snapshot_params()
            run_id = db.create_run(target_date, stable_days, volatility_threshold,
                                   conversion_rate, target_count)
            st.session_state.current_run_id = run_id
            progress_bar = st.progress(0)

            # Step 1
            st.session_state.stage2_done = False
            st.session_state.stable_items = []
            st.session_state.stage3_done = False
            st.session_state.steam_data = {}
            st.session_state.stage4_done = False
            st.session_state.arbitrage_results = {}
            with st.status("第1步/共4步：BUFF初步筛选…", expanded=True) as status:
                categories_label = '、'.join(category_names) if category_names else '全部/不限'
                st.write("### 正在从 BUFF 市场抓取饰品列表…")
                st.write(f"**筛选条件:**")
                st.write(f"  · 品类: {categories_label}")
                st.write(f"  · 目标数量: {target_count} 件")
                st.write(f"  · 最低价格: ¥{min_price}")
                st.write(f"  · 最低在售: {min_volume} 件")
                _execute_step1(target_date, target_count, category_values,
                               min_price, min_volume, run_id)
                raw = st.session_state.raw_items
                filtered = st.session_state.filtered_items
                st.write(f"---")
                st.write(f"**抓取结果:**")
                st.write(f"  · BUFF 原始获取: {len(raw)} 条")
                st.write(f"  · 初步过滤（在售>{min_volume}，价格>¥{min_price}）: {len(filtered)} 条")
                if filtered:
                    st.write(f"**已获取饰品列表:**")
                    for it in filtered[:20]:
                        st.write(f"  · {it.name}: ¥{it.buff_price:.2f}, 在售 {it.volume} 件")
                    if len(filtered) > 20:
                        st.write(f"  ... 及 {len(filtered)-20} 件更多")
                status.update(label=f"✅ 第1步完成：BUFF初步筛选 — {len(filtered)} 条通过", state="complete")
            progress_bar.progress(0.25)

            # Step 2
            with st.status("第2步/共4步：价格稳定性筛选…", expanded=True) as status:
                if st.session_state.filtered_items:
                    st.write(f"### 正在对 {len(st.session_state.filtered_items)} 个饰品逐个检查价格稳定性…")
                    st.write(f"**参数:** 考察 {stable_days} 天, 波动阈值 ≤ {volatility_threshold*100:.0f}%")
                    st.write(f"---")
                    _execute_step2(target_date, stable_days, volatility_threshold, run_id)
                stable = st.session_state.stable_items
                st.write(f"---")
                st.write(f"**结果汇总:**")
                st.write(f"  · 通过（波动 ≤ {volatility_threshold*100:.0f}%）: {len(stable)} 条")
                fail_count = len(st.session_state.filtered_items) - len(stable)
                st.write(f"  · 未通过: {fail_count} 条")
                if stable:
                    st.write(f"**已筛选通过的稳定饰品:**")
                    for item in stable:
                        st.write(f"  · {item.name}: ¥{item.buff_price:.2f}, "
                                 f"波动 {item._debug_volatility*100:.1f}%")
                status.update(label=f"✅ 第2步完成：价格稳定性筛选 — {len(stable)} 条通过", state="complete")
            progress_bar.progress(0.5)

            # Step 3
            with st.status("第3步/共4步：Steam市场数据获取…", expanded=True) as status:
                if st.session_state.stable_items:
                    st.write(f"### 正在对 {len(st.session_state.stable_items)} 个饰品抓取 Steam 市场数据…")
                    st.write(f"**说明:** 每个饰品会打开 Steam 市场页面，匹配对应磨损/StatTrak 变体，")
                    st.write(f"获取历史售价及已售数量")
                    st.write(f"---")
                    _execute_step3(run_id, status=status)
                steam_data = st.session_state.steam_data
                fail_count = len(stable) - len(steam_data)
                st.write(f"---")
                st.write(f"**结果汇总:**")
                st.write(f"  · 成功获取 Steam 数据: {len(steam_data)} / {len(stable)} 条")
                if fail_count > 0:
                    st.write(f"  · 获取失败: {fail_count} 条")
                if steam_data:
                    st.write(f"**已获取数据列表:**")
                    for sid, data in steam_data.items():
                        item = next((it for it in stable if it.item_id == sid), None)
                        name = item.name if item else sid
                        sp = data.get("steam_price")
                        sc = data.get("steam_sold_count", 0)
                        sp_str = f"${sp:.2f}" if sp else "N/A"
                        st.write(f"  · {name}: Steam {sp_str}, 售出 {sc} 件")
                status.update(label=f"✅ 第3步完成：Steam数据 — {len(steam_data)}/{len(stable)} 条", state="complete")
            progress_bar.progress(0.75)

            # Step 4
            with st.status("第4步/共4步：套利对比分析…", expanded=True) as status:
                stable = st.session_state.stable_items
                st.write(f"### 正在对 {len(stable)} 个饰品进行 BUFF vs Steam 套利对比…")
                st.write(f"**卡价转换比:** 1 USD = {conversion_rate} CNY")
                st.write(f"**判定规则:** BUFF 均价 > Steam 转换均价 = 目标饰品")
                st.write(f"---")
                if st.session_state.steam_data:
                    _execute_step4(conversion_rate, run_id)
                arb = st.session_state.arbitrage_results
                found = sum(1 for v in arb.values() if v.get("is_target"))
                st.write(f"---")
                st.write(f"**结果汇总:**")
                st.write(f"  · 有可对比数据的饰品: {len(arb)} 条")
                st.write(f"  · 目标饰品（BUFF均价 > Steam转换均价）: {found} 个")
                if arb:
                    st.write(f"**各饰品对比详情:**")
                    for sid, ar_data in arb.items():
                        item = next((it for it in st.session_state.stable_items if it.item_id == sid), None)
                        name = item.name if item else sid
                        label = "🎯 目标" if ar_data.get("is_target") else "未达标"
                        st.write(f"  · {name}: BUFF ¥{ar_data['avg_buff_price']:.2f} vs "
                                 f"Steam ¥{ar_data['avg_steam_cny']:.2f}, "
                                 f"差¥{ar_data['avg_diff']:+.2f} — {label}")
                        for p in ar_data.get("date_pairs", []):
                            icon = "✅" if p["is_target"] else "❌"
                            st.write(f"      {icon} {p['buff_date']}: BUFF ¥{p['buff_price']:.2f} "
                                     f"vs Steam ${p['steam_price_usd']:.2f}→¥{p['steam_price_cny']:.2f} "
                                     f"(差¥{p['diff']:+.2f})")
                status.update(label=f"✅ 第4步完成：套利对比 — {found} 个目标", state="complete")
            progress_bar.progress(1.0)

            progress_bar.empty()
            st.session_state.one_click_mode = "done"
            st.rerun()

        st.divider()

        # ---- 分步执行 ----
        with st.expander("⚙️ 分步执行（含重新执行按钮）", expanded=False):
            stage_names = ["BUFF初步筛选", "价格稳定性筛选", "Steam市场数据获取", "套利对比"]
            done_count = sum([
                st.session_state.stage1_done,
                st.session_state.stage2_done,
                st.session_state.stage3_done,
                st.session_state.stage4_done,
            ])
            st.caption(f"当前进度：{done_count}/4  —  {stage_names[done_count] if done_count < 4 else '全部完成'}")

            # =====================================================================
            # 第一步：BUFF初步筛选
            # =====================================================================
            st.subheader("第一步：BUFF初步筛选")
            st.caption("从BUFF市场列表页抓取饰品（价格、在售数量），应用 在售>100 & 价格>20 初步过滤。")

            col1a, col1b, col1c = st.columns([1, 1, 2])
            with col1a:
                if st.button("执行第一步", type="primary",
                             disabled=st.session_state.stage1_done,
                             use_container_width=True,
                             key="step1_exec"):
                    st.session_state.stage2_done = False
                    st.session_state.stable_items = []
                    st.session_state.stage3_done = False
                    st.session_state.steam_data = {}
                    st.session_state.stage4_done = False
                    st.session_state.arbitrage_results = {}
                    st.session_state.error_log = []
                    st.session_state.one_click_mode = None
                    category_values = [config.CATEGORY_OPTIONS[n] for n in category_names if n != "全部/不限"]
                    _snapshot_params()
                    run_id = db.create_run(target_date, stable_days, volatility_threshold,
                                           conversion_rate, target_count)
                    st.session_state.current_run_id = run_id
                    with st.spinner("正在从BUFF获取饰品列表..."):
                        _execute_step1(target_date, target_count, category_values,
                                       min_price, min_volume, run_id)
                    st.rerun()
            with col1b:
                if st.button("🔄 重新执行", type="secondary",
                             disabled=not st.session_state.stage1_done,
                             use_container_width=True,
                             key="step1_redo"):
                    st.session_state.stage1_done = False
                    st.session_state.raw_items = []
                    st.session_state.filtered_items = []
                    st.session_state.stage2_done = False
                    st.session_state.stable_items = []
                    st.session_state.stage3_done = False
                    st.session_state.steam_data = {}
                    st.session_state.stage4_done = False
                    st.session_state.arbitrage_results = {}
                    st.session_state.error_log = []
                    st.session_state.one_click_mode = None
                    category_values = [config.CATEGORY_OPTIONS[n] for n in category_names if n != "全部/不限"]
                    _snapshot_params()
                    run_id = db.create_run(target_date, stable_days, volatility_threshold,
                                           conversion_rate, target_count)
                    st.session_state.current_run_id = run_id
                    with st.spinner("正在重新执行第一步..."):
                        _execute_step1(target_date, target_count, category_values,
                                       min_price, min_volume, run_id)
                    st.rerun()

            if st.session_state.stage1_done:
                raw = st.session_state.raw_items
                filtered = st.session_state.filtered_items
                st.write(f"BUFF 原始获取：{len(raw)} 条 → 初步过滤后：{len(filtered)} 条")

                if not filtered:
                    st.warning("没有饰品通过初步筛选（在售>100 且 价格>20）。请检查数据或放宽条件。")
                else:
                    df1 = pd.DataFrame([{
                        "饰品ID": it.item_id, "名称": it.name,
                        "BUFF价格": it.buff_price, "在售数量": it.volume, "成交额": it.turnover,
                    } for it in filtered])
                    st.dataframe(df1, use_container_width=True)

                # ---- 逐步诊断 ----
                if filtered:
                    with st.expander("🔍 逐步诊断：选择单个饰品，逐步检查价格提取的每步操作"):
                        st.markdown("选中一个饰品后点击按钮，下方会逐步展示整个提取流程中每一步的成功/失败详情。")
                        col_a, col_b = st.columns([2, 1])
                        with col_a:
                            test_idx = st.selectbox(
                                "选择饰品", range(len(filtered)),
                                format_func=lambda i: f"{filtered[i].name} (ID:{filtered[i].item_id})",
                                key="diag_select"
                            )
                        with col_b:
                            run_diag = st.button("🔬 开始逐步诊断", type="primary", use_container_width=True)

                        if run_diag:
                            item = filtered[test_idx]
                            start = target_date - timedelta(days=stable_days)
                            with st.spinner("正在逐步诊断价格提取..."):
                                diag = diagnose_price_extraction(item.item_id, start, target_date)

                            st.divider()
                            st.markdown(f"### 诊断结果：{item.name}")
                            st.caption(f"ID: {item.item_id} | 范围: {start} ~ {target_date} | 时间跨度: {diag['range_text']}")

                            ok_count = sum(1 for s in diag["steps"] if s["ok"])
                            total = len(diag["steps"])
                            if ok_count == total:
                                st.success(f"全部 {total} 步通过！")
                            else:
                                st.error(f"{ok_count}/{total} 步通过，{total - ok_count} 步失败")

                            for step in diag["steps"]:
                                icon = "✅" if step["ok"] else "❌"
                                with st.expander(f"{icon} {step['name']} — {step['detail'][:80]}", expanded=not step["ok"]):
                                    st.text(step["detail"])
                                    if step.get("screenshot"):
                                        st.caption(f"截图: {step['screenshot']}")
                                    if "data" in step and step["data"]:
                                        st.caption(f"Playwright 拦截到 {len(step['data'])} 个响应")
                                        for r in step["data"][:5]:
                                            url = r.get("url", "")[:200]
                                            ct = r.get("content_type", "")
                                            has_body = "body" in r
                                            st.code(f"[{r.get('status','?')}] {url}  ({ct}) {'JSON✅' if has_body else 'no body'}")
                                    if step.get("js_hooked"):
                                        st.caption(f"JS hook 捕获 {len(step['js_hooked'])} 条请求")
                                        seen = set()
                                        for entry in step["js_hooked"]:
                                            url = entry.get("url", "")
                                            if url and url not in seen:
                                                seen.add(url)
                                                st.code(f"[{entry.get('type','?')}] {entry.get('method','?')} {url[:200]}  status={entry.get('status','?')}")
                                    if step.get("all_hooked"):
                                        st.caption(f"最终 JS hook 捕获 {len(step['all_hooked'])} 条请求（含响应内容）")

                            st.divider()
                            st.subheader("最终提取结果")
                            if diag["records"]:
                                st.success(f"成功提取 {len(diag['records'])} 条价格记录")
                                df_diag = pd.DataFrame([{
                                    "日期": r.date, "价格": r.price
                                } for r in diag["records"]])
                                st.dataframe(df_diag, use_container_width=True)
                                prices = [r.price for r in diag["records"]]
                                st.write(f"最低: {min(prices):.2f} | 最高: {max(prices):.2f} | 平均: {sum(prices)/len(prices):.2f}")
                            else:
                                st.error("未能提取到任何价格记录")

                            st.caption(f"临时文件目录: {diag['tmp_dir']}")

            st.divider()

            # =====================================================================
            # 第二步：价格稳定性筛选
            # =====================================================================
            st.subheader("第二步：价格稳定性筛选")
            st.caption(f"对每个通过第一步的饰品，获取 {stable_days} 天内的 BUFF 价格历史，筛掉波动超过 {volatility_threshold * 100:.0f}% 的。")

            if st.session_state.stage1_done and st.session_state.filtered_items:
                col2a, col2b, col2c = st.columns([1, 1, 2])
                with col2a:
                    if st.button("执行第二步", type="primary",
                                 disabled=st.session_state.stage2_done,
                                 use_container_width=True,
                                 key="step2_exec"):
                        st.session_state.stage3_done = False
                        st.session_state.steam_data = {}
                        st.session_state.stage4_done = False
                        st.session_state.arbitrage_results = {}
                        st.session_state.one_click_mode = None
                        with st.spinner("正在执行价格稳定性筛选..."):
                            _execute_step2(target_date, stable_days, volatility_threshold,
                                           st.session_state.current_run_id)
                        st.rerun()
                with col2b:
                    if st.button("🔄 重新执行", type="secondary",
                                 disabled=not st.session_state.stage2_done,
                                 use_container_width=True,
                                 key="step2_redo"):
                        st.session_state.stage2_done = False
                        st.session_state.stable_items = []
                        st.session_state.stage3_done = False
                        st.session_state.steam_data = {}
                        st.session_state.stage4_done = False
                        st.session_state.arbitrage_results = {}
                        st.session_state.one_click_mode = None
                        with st.spinner("正在重新执行价格稳定性筛选..."):
                            _execute_step2(target_date, stable_days, volatility_threshold,
                                           st.session_state.current_run_id)
                        st.rerun()

            if st.session_state.stage2_done:
                stable = st.session_state.stable_items
                filtered = st.session_state.filtered_items
                st.write(f"价格稳定（波动 ≤ {volatility_threshold * 100:.0f}%）的饰品：{len(stable)} / {len(filtered)} 条")

                with st.expander("查看每条处理明细"):
                    rows = []
                    for item in filtered:
                        hist_len = getattr(item, '_debug_history_len', -1)
                        min_p = getattr(item, '_debug_min_price', 0)
                        max_p = getattr(item, '_debug_max_price', 0)
                        vol = getattr(item, '_debug_volatility', 0)
                        passed = item in stable
                        fail_reason = getattr(item, '_debug_fail_reason', '')
                        rows.append({
                            "饰品": item.name,
                            "价格记录数": hist_len,
                            "最低": f"{min_p:.2f}" if hist_len > 0 else "N/A",
                            "最高": f"{max_p:.2f}" if hist_len > 0 else "N/A",
                            "波动": f"{vol*100:.1f}%" if hist_len > 0 else "N/A",
                            "结果": "✅" if passed else f"❌ {fail_reason}",
                        })
                    st.dataframe(pd.DataFrame(rows), use_container_width=True)

                if not stable:
                    st.warning("没有饰品通过价格稳定性筛选，请放宽阈值或缩短考察天数后重新执行第二步。")

            # =====================================================================
            # 第三步：获取Steam市场价格与销量
            # =====================================================================
            st.divider()
            st.subheader("第三步：获取Steam市场价格与销量")
            st.caption('对每个通过价格稳定性筛选的饰品，打开浏览器点击"查看Steam市场"按钮获取Steam售价及已售数量。'
                       '每个日期点为BUFF价格历史日期-7天，自动匹配饰品变体（磨损/StatTrak），货币为美元。')

            if st.session_state.stage2_done and st.session_state.stable_items:
                col3a, col3b, col3c = st.columns([1, 1, 2])
                with col3a:
                    if st.button("执行第三步", type="primary",
                                 disabled=st.session_state.stage3_done,
                                 use_container_width=True,
                                 key="step3_exec"):
                        st.session_state.stage4_done = False
                        st.session_state.arbitrage_results = {}
                        st.session_state.one_click_mode = None
                        with st.spinner("正在获取Steam市场数据..."):
                            _execute_step3(st.session_state.current_run_id)
                        st.rerun()
                with col3b:
                    if st.button("🔄 重新执行", type="secondary",
                                 disabled=not st.session_state.stage3_done,
                                 use_container_width=True,
                                 key="step3_redo"):
                        st.session_state.stage3_done = False
                        st.session_state.steam_data = {}
                        st.session_state.stage4_done = False
                        st.session_state.arbitrage_results = {}
                        st.session_state.one_click_mode = None
                        with st.spinner("正在重新获取Steam市场数据..."):
                            _execute_step3(st.session_state.current_run_id)
                        st.rerun()

            if st.session_state.stage3_done:
                steam_data = st.session_state.steam_data
                stable = st.session_state.stable_items
                success_count = len(steam_data)
                failed_items = st.session_state.get("failed_step3_items", [])
                fail_count = len(stable) - success_count
                st.write(f"成功获取Steam市场数据：{success_count} / {len(stable)} 条")
                if fail_count > 0:
                    st.write(f"获取失败: {fail_count} 条")

                # ── 重试失败饰品 ──
                if failed_items:
                    with st.expander(f"🔄 重试失败饰品（{len(failed_items)} 条）", expanded=False):
                        st.markdown("以下饰品 Step 3 获取失败，可单独或批量重试：")
                        for fi in failed_items:
                            name = fi["buff_item_name"]
                            col_a, col_b = st.columns([4, 1])
                            with col_a:
                                st.caption(f"{name}  (ID: {fi['item_id']})")
                            with col_b:
                                retry_key = f"retry_{fi['item_id']}"
                                if st.button("重试", key=retry_key, use_container_width=True):
                                    with st.spinner(f"正在重试 {name}..."):
                                        result = _steam.get_steam_market_data(
                                            fi["item_id"],
                                            fi["target_dates"],
                                            fi["buff_item_name"],
                                        )
                                    if result:
                                        steam_data[fi["item_id"]] = result
                                        st.success(f"✅ {name} 重试成功")
                                        st.session_state.failed_step3_items = [
                                            f for f in st.session_state.failed_step3_items
                                            if f["item_id"] != fi["item_id"]
                                        ]
                                        st.rerun()
                                    else:
                                        err = _steam.get_last_steam_error() or "未知错误"
                                        ctx = _steam.get_last_steam_error_context()
                                        _log_error(3, fi["item_id"], name, err, context=ctx)
                                        st.error(f"❌ {name} 重试失败: {err}")

                        st.divider()
                        if st.button("🔄 重试全部失败饰品", type="primary",
                                      use_container_width=True, key="retry_all_failed"):
                            with st.status("正在批量重试所有失败饰品...",
                                           expanded=True) as retry_status:
                                batch_result = _steam.retry_steam_failed_items(failed_items)
                                ok = 0
                                for fi in failed_items:
                                    data = batch_result.get(fi["item_id"])
                                    if data:
                                        steam_data[fi["item_id"]] = data
                                        ok += 1
                                        retry_status.write(f"✅ {fi['buff_item_name']} 成功")
                                    else:
                                        err = _steam.get_last_steam_error(item_id=fi["item_id"]) or "失败"
                                        _log_error(3, fi["item_id"], fi["buff_item_name"], err)
                                        retry_status.write(f"❌ {fi['buff_item_name']} 失败")
                                st.session_state.failed_step3_items = [
                                    f for f in failed_items
                                    if f["item_id"] not in batch_result
                                ]
                                st.success(f"重试完成: {ok}/{len(failed_items)} 成功")
                            st.rerun()

                if steam_data:
                    rows = []
                    for item in stable:
                        data = steam_data.get(item.item_id)
                        if data:
                            date_records = data.get("date_records", [])
                            date_details = "; ".join(
                                f"{d['date']}: ${d['steam_price']:.2f}/{d['steam_volume']}件"
                                for d in date_records
                            ) if date_records else "无数据"
                            sp = data.get("steam_price")
                            sc = data.get("steam_sold_count", 0)
                            steam_price_str = f"${sp:.2f}" if sp else "N/A"
                            rows.append({
                                "饰品ID": item.item_id,
                                "名称": item.name,
                                "BUFF价格": item.buff_price,
                                "Steam均价": steam_price_str,
                                "总售出量": sc,
                                "各日期节点明细": date_details,
                                "Steam链接": data.get("steam_url", "获取失败")[:80] + "...",
                            })
                        else:
                            rows.append({
                                "饰品ID": item.item_id, "名称": item.name,
                                "BUFF价格": item.buff_price, "Steam均价": "获取失败",
                                "总售出量": 0, "各日期节点明细": "获取失败",
                                "Steam链接": "获取失败",
                            })
                    st.dataframe(pd.DataFrame(rows), use_container_width=True)

                    with st.expander("📊 按日期节点查看 Steam 价格 & 销量明细"):
                        for item in stable:
                            data = steam_data.get(item.item_id)
                            if data and data.get("date_records"):
                                st.markdown(f"**{item.name}** (ID: {item.item_id})")
                                df_dates = pd.DataFrame(data["date_records"])
                                df_dates.columns = ["日期", "Steam价格($)", "Steam销量(件)"]
                                st.dataframe(df_dates, use_container_width=True)
                                st.divider()
                            elif data:
                                st.markdown(f"**{item.name}** — 无日期记录")

                    with st.expander("点击查看Steam市场链接"):
                        for item in stable:
                            data = steam_data.get(item.item_id)
                            url = data.get("steam_url") if data else None
                            if url:
                                st.markdown(f"- [{item.name} (ID:{item.item_id})]({url})")
                            else:
                                st.markdown(f"- {item.name} (ID:{item.item_id}): 获取失败")
                else:
                    st.warning("未能获取到任何Steam市场数据。")

                # ---- Step 3 逐步诊断 ----
                if steam_data:
                    with st.expander("🔍 逐步诊断：Steam数据提取流程", expanded=False):
                        st.markdown("选择饰品后点击按钮，逐步检查 Steam 数据提取的每步操作。")
                        col_a, col_b = st.columns([2, 1])
                        with col_a:
                            diag_items = [it for it in stable if steam_data.get(it.item_id)]
                            if diag_items:
                                test_idx = st.selectbox(
                                    "选择饰品", range(len(diag_items)),
                                    format_func=lambda i: f"{diag_items[i].name} (ID:{diag_items[i].item_id})",
                                    key="steam_diag_select"
                                )
                            else:
                                test_idx = None
                        with col_b:
                            run_diag = st.button("🔬 开始Steam诊断", type="primary", use_container_width=True,
                                                 key="steam_diag_btn")

                        if run_diag and diag_items:
                            item = diag_items[test_idx]
                            target_dates = sorted(set(
                                r.date - timedelta(days=7) for r in item.price_history
                            ))
                            with st.spinner("正在逐步诊断 Steam 数据提取..."):
                                diag = diagnose_steam_extraction(
                                    item.item_id, target_dates, item.name
                                )

                            st.divider()
                            st.markdown(f"### Steam诊断结果：{item.name}")
                            st.caption(f"ID: {item.item_id} | "
                                       f"目标日期: {[str(d) for d in target_dates]}")

                            ok_count = sum(1 for s in diag["steps"] if s["ok"])
                            total = len(diag["steps"])
                            if ok_count == total:
                                st.success(f"全部 {total} 步通过！")
                            else:
                                st.error(f"{ok_count}/{total} 步通过，{total - ok_count} 步失败")

                            for step in diag["steps"]:
                                icon = "✅" if step["ok"] else "❌"
                                with st.expander(f"{icon} {step['name']} — {step['detail'][:80]}",
                                                 expanded=not step["ok"]):
                                    st.text(step["detail"])
                                    if step.get("screenshot"):
                                        st.caption(f"截图: {step['screenshot']}")
                                    if step.get("data"):
                                        st.json(step["data"])

                            st.divider()
                            st.subheader("提取的价格记录")
                            if diag["price_history"]:
                                st.success(f"成功提取 {len(diag['price_history'])} 条记录")
                                df_diag = pd.DataFrame([{
                                    "日期": r.date,
                                    "Steam价格($)": f"{r.price:.2f}",
                                    "销量(件)": r.volume,
                                } for r in diag["price_history"]])
                                st.dataframe(df_diag, use_container_width=True)
                            else:
                                st.error("未能提取到任何价格记录")

                            st.caption(f"临时文件目录: {diag['tmp_dir']}")

            # =====================================================================
            # 第四步：BUFF vs Steam 价格对比 & 套利筛选
            # =====================================================================
            st.divider()
            st.subheader("第四步：BUFF vs Steam 价格对比")
            st.caption('对每个饰品按日期节点配对，将Steam价格乘上卡价转换比转换为人民币，与BUFF价格对比。'
                       '当BUFF均价高于Steam转换均价时，标记为目标饰品。')

            if st.session_state.stage3_done and st.session_state.steam_data:
                col4a, col4b, col4c = st.columns([1, 1, 2])
                with col4a:
                    if st.button("执行第四步", type="primary",
                                 disabled=st.session_state.stage4_done,
                                 use_container_width=True,
                                 key="step4_exec"):
                        st.session_state.one_click_mode = None
                        with st.spinner("正在执行套利对比..."):
                            _execute_step4(conversion_rate, st.session_state.current_run_id)
                        st.rerun()
                with col4b:
                    if st.button("🔄 重新执行", type="secondary",
                                 disabled=not st.session_state.stage4_done,
                                 use_container_width=True,
                                 key="step4_redo"):
                        st.session_state.stage4_done = False
                        st.session_state.arbitrage_results = {}
                        st.session_state.one_click_mode = None
                        with st.spinner("正在重新执行套利对比..."):
                            _execute_step4(conversion_rate, st.session_state.current_run_id)
                        st.rerun()

            if st.session_state.stage4_done and not one_click_mode == "done":
                arbitrage_results = st.session_state.arbitrage_results
                stable = st.session_state.stable_items
                steam_data = st.session_state.steam_data

                if arbitrage_results:
                    target_items = [it for it in stable if arbitrage_results.get(it.item_id, {}).get("is_target")]
                    st.write(f"套利分析完成：{len(arbitrage_results)} 个饰品有对比数据，"
                             f"其中 **{len(target_items)} 个目标饰品**")

                    rows = []
                    for item in stable:
                        ar = arbitrage_results.get(item.item_id)
                        if ar:
                            date_details = "; ".join(
                                f"{p['buff_date']}: BUFF ¥{p['buff_price']:.2f} vs "
                                f"Steam ${p['steam_price_usd']:.2f}×{conversion_rate}=¥{p['steam_price_cny']:.2f} "
                                f"→ {'✅' if p['is_target'] else '❌'}差¥{p['diff']:+.2f}"
                                for p in ar["date_pairs"]
                            )
                            rows.append({
                                "饰品ID": item.item_id, "名称": item.name,
                                "BUFF均价(¥)": f"{ar['avg_buff_price']:.2f}",
                                "Steam均价($)": f"${ar['avg_steam_usd']:.2f}",
                                "Steam均价(¥)": f"¥{ar['avg_steam_cny']:.2f}",
                                "均价差(¥)": f"¥{ar['avg_diff']:+.2f}",
                                "命中节点": f"{ar['target_count']}/{len(ar['date_pairs'])}",
                                "判定": "🎯 目标" if ar["is_target"] else "未达标",
                                "各节点明细": date_details,
                            })
                        else:
                            rows.append({
                                "饰品ID": item.item_id, "名称": item.name,
                                "BUFF均价(¥)": item.buff_price, "Steam均价($)": "无数据",
                                "Steam均价(¥)": "无数据", "均价差(¥)": "N/A",
                                "命中节点": "N/A", "判定": "无Steam数据", "各节点明细": "无",
                            })
                    st.dataframe(pd.DataFrame(rows), use_container_width=True)

                    with st.expander("📋 查看每条处理明细", expanded=False):
                        detail_rows = []
                        for item in stable:
                            ar = arbitrage_results.get(item.item_id)
                            if not ar:
                                detail_rows.append({
                                    "饰品": item.name, "对比节点数": 0,
                                    "BUFF均价": f"{item.buff_price:.2f}",
                                    "Steam转换均价": "N/A", "均价差": "N/A",
                                    "命中/总节点": "N/A", "结果": "❌ 无Steam数据",
                                })
                                continue
                            detail_rows.append({
                                "饰品": item.name, "对比节点数": len(ar["date_pairs"]),
                                "BUFF均价": f"¥{ar['avg_buff_price']:.2f}",
                                "Steam转换均价": f"¥{ar['avg_steam_cny']:.2f}",
                                "均价差": f"¥{ar['avg_diff']:+.2f}",
                                "命中/总节点": f"{ar['target_count']}/{len(ar['date_pairs'])}",
                                "结果": "🎯 目标" if ar["is_target"] else "未达标",
                            })
                        st.dataframe(pd.DataFrame(detail_rows), use_container_width=True)

                    with st.expander("📊 按日期节点查看配对对比"):
                        for item in stable:
                            ar = arbitrage_results.get(item.item_id)
                            if ar and ar["date_pairs"]:
                                st.markdown(f"**{item.name}** (ID: {item.item_id})  "
                                            f"{'🎯 目标' if ar['is_target'] else '未达标'}")
                                pair_rows = []
                                for p in ar["date_pairs"]:
                                    pair_rows.append({
                                        "BUFF日期": p["buff_date"], "BUFF价格(¥)": f"{p['buff_price']:.2f}",
                                        "Steam日期": p["steam_date"], "Steam价格($)": f"{p['steam_price_usd']:.2f}",
                                        "Steam转换(¥)": f"{p['steam_price_cny']:.2f}",
                                        "差价(¥)": f"{p['diff']:+.2f}", "Steam销量": p["steam_volume"],
                                        "判定": "✅ 目标" if p["is_target"] else "❌",
                                    })
                                st.dataframe(pd.DataFrame(pair_rows), use_container_width=True)
                                st.divider()

                    with st.expander("📈 BUFF vs Steam 价格对比图表"):
                        seen = {}
                        chart_idx = 0
                        for item in stable:
                            if item.item_id in seen:
                                continue
                            seen[item.item_id] = True
                            data = steam_data.get(item.item_id)
                            ar = arbitrage_results.get(item.item_id)
                            if data and data.get("steam_price_history"):
                                try:
                                    fig = go.Figure()
                                    if item.price_history:
                                        buff_dates = [r.date for r in item.price_history]
                                        buff_prices = [r.price for r in item.price_history]
                                        fig.add_trace(go.Scatter(
                                            x=buff_dates, y=buff_prices, name=f"{item.name} BUFF(¥)",
                                            mode="lines+markers", yaxis="y",
                                        ))
                                    steam_hist = data["steam_price_history"]
                                    steam_dates = [r.date for r in steam_hist]
                                    steam_prices_cny = [r.price * conversion_rate for r in steam_hist]
                                    fig.add_trace(go.Scatter(
                                        x=steam_dates, y=steam_prices_cny, name=f"{item.name} Steam→¥",
                                        mode="lines+markers", yaxis="y",
                                    ))
                                    title_suffix = " 🎯目标" if (ar and ar["is_target"]) else ""
                                    fig.update_layout(
                                        title=f"{item.name} 价格对比 (1 USD = {conversion_rate} CNY){title_suffix}",
                                        xaxis_title="日期", yaxis_title="价格 (¥)",
                                        hovermode="x unified",
                                    )
                                    st.plotly_chart(fig, use_container_width=True,
                                                    key=f"step4_chart_{chart_idx}")
                                    chart_idx += 1
                                except ImportError:
                                    st.info("安装 plotly 可查看价格对比图表：pip install plotly")
                                    break

                else:
                    st.warning("未能进行套利对比分析，请确认 Step 3 返回了有效数据。")

            # ---- 全局重置按钮（仅分步模式显示） ----
            if any([st.session_state.stage1_done, st.session_state.stage2_done,
                    st.session_state.stage3_done, st.session_state.stage4_done]):
                _show_error_log()
                st.divider()
                if st.button("🔄 重新开始（清除所有结果）", use_container_width=True, key="reset_all"):
                    for k in defaults:
                        st.session_state[k] = defaults[k]
                    st.rerun()

with tabs[1]:
    st.subheader("BUFF价格走势查询")

    col1, col2 = st.columns([2, 1])
    with col1:
        item_id_input = st.text_input(
            "请输入饰品ID（可从BUFF商品链接中获取）",
            value="",
            key="buff_price_trend_item_id",
        )
    with col2:
        time_range = st.selectbox(
            "日期范围",
            options=["最近3个月", "最近6个月", "最近1年"],
            index=0,
            key="buff_price_trend_time_range",
        )

    if st.button("查询BUFF价格走势", type="primary", key="query_buff_price_trend"):
        if not item_id_input.strip():
            st.warning("请输入饰品ID")
        else:
            with st.spinner("正在从BUFF获取价格走势数据..."):
                history_data = get_full_price_history(item_id_input.strip(), time_range)

            if not history_data:
                st.error("未能获取到价格走势数据，请检查饰品ID是否正确或BUFF登录状态。")
            else:
                st.success(f"成功获取 {len(history_data)} 条曲线数据")

                try:

                    fig = go.Figure()

                    preferred_order = [
                        "在售最低", "求购最高", "在售数量", "成交记录", "买卖点", "存世量",
                    ]
                    sorted_names = []
                    for name in preferred_order:
                        if name in history_data:
                            sorted_names.append(name)
                    for name in history_data:
                        if name not in sorted_names:
                            sorted_names.append(name)

                    for name in sorted_names:
                        data = history_data[name]
                        dates = [r[0] for r in data]
                        values = [r[1] for r in data]

                        visible = True if name == "在售最低" else "legendonly"

                        fig.add_trace(
                            go.Scatter(
                                x=dates, y=values, name=name,
                                mode="lines", visible=visible,
                                hovertemplate=f"{name}: %{{y:.2f}}<extra></extra>",
                            )
                        )

                    fig.update_layout(
                        hovermode="x unified",
                        xaxis_title="日期",
                        yaxis_title="价格",
                        legend=dict(
                            orientation="h", yanchor="bottom",
                            y=1.02, xanchor="center", x=0.5,
                        ),
                        margin=dict(t=80),
                    )

                    st.plotly_chart(fig, use_container_width=True,
                                    key="buff_price_trend_chart")
                except ImportError:
                    st.error("请安装 plotly 以查看交互式图表：pip install plotly")
